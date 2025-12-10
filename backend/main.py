from fastapi import FastAPI, Query, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy import create_engine, text
from pydantic import BaseModel
import os
import shutil
import json
import re
import gc
import traceback
import time
from datetime import datetime
from dotenv import load_dotenv
from cryptography.fernet import Fernet
import openpyxl

# 1. Configuration
load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL")
ENCRYPTION_KEY = os.getenv("ENCRYPTION_KEY", Fernet.generate_key().decode())
cipher = Fernet(ENCRYPTION_KEY.encode())

app = FastAPI()

# 2. Sécurité
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_methods=["*"],
    allow_headers=["*"],
)

# 3. Connexion BDD Robuste
engine = None
if DATABASE_URL:
    if DATABASE_URL.startswith("postgres://"):
        DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
    
    if "sslmode" not in DATABASE_URL:
        separator = "&" if "?" in DATABASE_URL else "?"
        DATABASE_URL += f"{separator}sslmode=require"

    try:
        # pool_pre_ping=True est vital pour la stabilité sur le cloud
        engine = create_engine(DATABASE_URL, pool_pre_ping=True, pool_recycle=300)
        with engine.begin() as conn:
            # --- Table Utilisateurs (Auth) ---
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS users (
                    username VARCHAR(100) PRIMARY KEY,
                    shop_name TEXT,
                    password TEXT,
                    email TEXT,
                    role VARCHAR(50) DEFAULT 'user',
                    is_first_login BOOLEAN DEFAULT TRUE
                );
            """))
            
            # Initialisation Admin par défaut
            try:
                user_count = conn.execute(text("SELECT COUNT(*) FROM users")).scalar()
                if user_count == 0:
                    print("⚠️ Base utilisateurs vide : Création de l'admin par défaut...")
                    conn.execute(text("""
                        INSERT INTO users (username, shop_name, password, email, role, is_first_login) 
                        VALUES ('admin', 'ADMINISTRATION', 'admin', 'admin@podium.optique', 'admin', FALSE)
                    """))
                    print("✅ Utilisateur 'admin' créé (Mdp: admin)")
            except Exception as e:
                print(f"⚠️ Info Admin Init: {e}")

            # --- Table Dossiers Clients ---
            # Inclut les colonnes pour l'hyperviseur (username, tags)
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS client_offers (
                    id SERIAL PRIMARY KEY, 
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    username VARCHAR(100), -- Lien vers l'opticien
                    encrypted_identity TEXT, 
                    lens_details JSONB, 
                    financials JSONB,
                    tags JSONB -- Stockage structuré pour stats
                );
            """))
            
            # Migration automatique : Ajout colonnes si manquantes (pour compatibilité)
            try:
                conn.execute(text("ALTER TABLE client_offers ADD COLUMN IF NOT EXISTS username VARCHAR(100);"))
                conn.execute(text("ALTER TABLE client_offers ADD COLUMN IF NOT EXISTS tags JSONB;"))
            except Exception as e: print(f"Info Migration Client Offers: {e}")

            # --- Table Catalogue Verres ---
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS lenses (
                    id SERIAL PRIMARY KEY, brand TEXT, name TEXT, commercial_code TEXT,
                    geometry TEXT, design TEXT, index_mat TEXT, material TEXT, coating TEXT, 
                    commercial_flow TEXT, color TEXT, 
                    purchase_price DECIMAL(10,2), 
                    purchase_price_bonifie DECIMAL(10,2), -- Prix Alternance Bonifié
                    purchase_price_super_bonifie DECIMAL(10,2), -- Prix Alternance Super Bonifié
                    selling_price DECIMAL(10,2),
                    sell_kalixia DECIMAL(10,2), sell_itelis DECIMAL(10,2), 
                    sell_carteblanche DECIMAL(10,2), sell_seveane DECIMAL(10,2), 
                    sell_santeclair DECIMAL(10,2)
                );
            """))
            
            # Migration automatique : Ajout colonnes prix bonifiés si manquantes
            try:
                conn.execute(text("ALTER TABLE lenses ADD COLUMN IF NOT EXISTS purchase_price_bonifie DECIMAL(10,2) DEFAULT 0;"))
                conn.execute(text("ALTER TABLE lenses ADD COLUMN IF NOT EXISTS purchase_price_super_bonifie DECIMAL(10,2) DEFAULT 0;"))
            except Exception as e: print(f"Info Migration Lenses: {e}")

    except Exception as e:
        print(f"❌ ERREUR BDD STARTUP: {e}")
        engine = None

# --- OUTILS ---
def encrypt_dict(data): return cipher.encrypt(json.dumps(data).encode()).decode()
def decrypt_dict(token): 
    try: return json.loads(cipher.decrypt(token.encode()).decode())
    except: return {"name": "Donnée", "firstname": "Illisible", "dob": "?"}

def clean_price(value):
    if not value or value == '' or value == '-': return 0.0
    try:
        clean_str = str(value).replace('€', '').replace('â‚¬', '').replace('%', '').replace(' ', '').replace('\xa0', '').replace(',', '.')
        return float(clean_str)
    except: return 0.0

def clean_index(value):
    if not value: return "1.50"
    match = re.search(r"\d+\.?\d*", str(value).replace(',', '.'))
    return "{:.2f}".format(float(match.group(0))) if match else "1.50"

def clean_text(value): return str(value).strip() if value else ""

def normalize_string(text):
    if not text: return ""
    text = str(text).upper().strip()
    replacements = {'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E', 'À': 'A', 'Â': 'A', 'Î': 'I', 'Ï': 'I', 'Ô': 'O', 'Ù': 'U', 'Ç': 'C'}
    for acc, char in replacements.items(): text = text.replace(acc, char)
    return text

def get_col_idx(headers, candidates):
    for i, h in enumerate(headers):
        if h:
            h_str = normalize_string(h)
            for c in candidates:
                if normalize_string(c) in h_str: return i
    return -1

# --- MODELES ---
class LoginRequest(BaseModel): username: str; password: str
class PasswordUpdate(BaseModel): username: str; new_password: str
class OfferRequest(BaseModel): 
    username: str # Ajouté pour lier l'offre à l'opticien
    client: dict
    lens: dict
    finance: dict
    tags: dict # Ajouté pour l'hyperviseur

# --- ROUTES AUTHENTIFICATION ---
@app.post("/auth/login")
def login(creds: LoginRequest):
    if not engine: raise HTTPException(500, "Pas de BDD")
    try:
        with engine.connect() as conn:
            result = conn.execute(text("SELECT * FROM users WHERE LOWER(username) = LOWER(:u)"), {"u": creds.username}).fetchone()
            if not result: raise HTTPException(401, "Utilisateur inconnu")
            user = result._mapping
            if user['password'] != creds.password: raise HTTPException(401, "Mot de passe incorrect")
            
            role = user['role'] if 'role' in user else 'user'
            
            return {
                "status": "success",
                "user": {
                    "username": user['username'], "shop_name": user['shop_name'],
                    "email": user['email'], "role": role, "is_first_login": user['is_first_login']
                }
            }
    except Exception as e:
        if isinstance(e, HTTPException): raise e
        print(f"Login Error: {e}")
        raise HTTPException(500, str(e))

@app.post("/auth/update-password")
def update_password(data: PasswordUpdate):
    if not engine: raise HTTPException(500, "Pas de BDD")
    try:
        with engine.begin() as conn:
            conn.execute(text("UPDATE users SET password = :p, is_first_login = FALSE WHERE username = :u"), 
                         {"p": data.new_password, "u": data.username})
        return {"status": "success"}
    except Exception as e: raise HTTPException(500, str(e))

@app.post("/upload-users")
def upload_users(file: UploadFile = File(...)):
    print("🚀 Début upload USERS (Mode UPSERT)...", flush=True)
    if not engine: raise HTTPException(500, "Pas de BDD")
    
    temp = f"/tmp/users_{int(time.time())}.xlsx"
    try:
        with open(temp, "wb") as buffer: shutil.copyfileobj(file.file, buffer)
        
        print("📖 Lecture Excel Users...", flush=True)
        wb = openpyxl.load_workbook(temp, data_only=True)
        sheet = wb.active
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows: raise Exception("Fichier vide")

        # Scan En-têtes
        header_idx = -1
        headers = []
        for i, row in enumerate(rows[:20]):
            row_str = [str(c).upper() for c in row if c]
            if any(k in s for s in row_str for k in ["IDENTIFIANT", "USERNAME", "LOGIN", "ID CLIENT"]):
                headers, header_idx = row, i; break
        
        if header_idx == -1: headers, header_idx = rows[0], 0

        # Mapping
        c_id = get_col_idx(headers, ['IDENTIFIANT', 'ID', 'USERNAME', 'LOGIN'])
        c_shop = get_col_idx(headers, ['MAGASIN', 'SHOP', 'RAISON SOCIALE', 'NOM'])
        c_pass = get_col_idx(headers, ['PASSWORD', 'MOT DE PASSE', 'MDP'])
        c_mail = get_col_idx(headers, ['MAIL', 'EMAIL', 'COURRIEL'])
        c_role = get_col_idx(headers, ['ROLE', 'TYPE', 'DROIT'])
        
        if c_id == -1: raise Exception("Colonne 'IDENTIFIANT' introuvable.")

        users_to_insert = []
        for row in rows[header_idx+1:]:
            if not row[c_id]: continue 
            
            u_id = str(row[c_id]).strip()
            u_shop = str(row[c_shop]).strip() if c_shop != -1 and row[c_shop] else "Opticien"
            u_pass = str(row[c_pass]).strip() if c_pass != -1 and row[c_pass] else "1234"
            u_mail = str(row[c_mail]).strip() if c_mail != -1 and row[c_mail] else ""
            
            u_role = "user"
            if c_role != -1 and row[c_role]:
                val_role = str(row[c_role]).lower().strip()
                if "admin" in val_role: u_role = "admin"
            
            users_to_insert.append({ "u": u_id, "s": u_shop, "p": u_pass, "e": u_mail, "r": u_role })
            
        print(f"✅ {len(users_to_insert)} utilisateurs prêts pour traitement.", flush=True)
            
        if users_to_insert:
            with engine.begin() as conn:
                # UPSERT : On ne touche pas au mot de passe si l'utilisateur existe
                query = text("""
                    INSERT INTO users (username, shop_name, password, email, role, is_first_login)
                    VALUES (:u, :s, :p, :e, :r, TRUE)
                    ON CONFLICT (username) DO UPDATE SET
                        shop_name = EXCLUDED.shop_name,
                        email = EXCLUDED.email,
                        role = EXCLUDED.role;
                """)
                conn.execute(query, users_to_insert)
                
        return {"status": "success", "count": len(users_to_insert), "mode": "upsert_safe"}

    except Exception as e: 
        print(f"❌ ERREUR UPLOAD USERS: {traceback.format_exc()}", flush=True)
        raise HTTPException(500, f"Erreur technique: {str(e)}")
    finally:
        if os.path.exists(temp): os.remove(temp)

# --- ROUTES ADMIN / HYPERVISEUR ---
@app.get("/admin/users")
def get_all_users():
    if not engine: return []
    try:
        with engine.connect() as conn:
            res = conn.execute(text("SELECT username, shop_name FROM users ORDER BY shop_name"))
            return [dict(r._mapping) for r in res]
    except: return []

@app.get("/admin/stats")
def get_user_stats(username: str = Query(...)):
    if not engine: return {}
    try:
        with engine.connect() as conn:
            # On récupère tags ET détails financiers pour calculs avancés
            res = conn.execute(text("SELECT tags, lens_details, financials FROM client_offers WHERE username = :u AND tags IS NOT NULL"), {"u": username})
            rows = res.fetchall()
            
            # Structure de base pour les stats catégorielles (Volume & Valeur)
            categories = ["network", "geometry", "design", "index", "material", "coating", "commercial_flow"]
            stats = {cat: {} for cat in categories}
            
            # Structure pour les Top Designs
            top_data = { "UNIFOCAL": {}, "PROGRESSIF": {} }
            
            total_sales = 0
            total_revenue = 0.0
            
            for r in rows:
                tags = r.tags
                if not tags: continue
                
                # Extraction Financière
                try:
                    # Prix total dossier (2 verres)
                    price = float(r.financials.get('total', 0))
                    # Marge unitaire * 2
                    unit_sell = float(r.lens_details.get('sellingPrice', 0))
                    unit_buy = float(r.lens_details.get('purchase_price', 0))
                    margin = (unit_sell - unit_buy) * 2
                except:
                    price = 0.0
                    margin = 0.0

                total_sales += 1
                total_revenue += price
                
                # 1. Agrégation par Catégories
                for cat in categories:
                    val = tags.get(cat, "N/A") or "N/A"
                    if val not in stats[cat]:
                        stats[cat][val] = {"volume": 0, "value": 0.0}
                    
                    stats[cat][val]["volume"] += 1
                    stats[cat][val]["value"] += price
                
                # 2. Agrégation Top Designs
                geo_raw = str(tags.get('geometry', '')).upper()
                design_name = str(tags.get('design', 'INCONNU'))
                
                target_geo = None
                if 'UNIFOCAL' in geo_raw: target_geo = "UNIFOCAL"
                elif 'PROGRESSIF' in geo_raw or 'INTERIEUR' in geo_raw: target_geo = "PROGRESSIF"
                
                if target_geo:
                    if design_name not in top_data[target_geo]:
                        top_data[target_geo][design_name] = {"volume": 0, "value": 0.0, "margin": 0.0}
                    
                    top_data[target_geo][design_name]["volume"] += 1
                    top_data[target_geo][design_name]["value"] += price
                    top_data[target_geo][design_name]["margin"] += margin

            # Formatage des Tops pour le frontend (Listes triées)
            final_tops = {"UNIFOCAL": {}, "PROGRESSIF": {}}
            for geo in ["UNIFOCAL", "PROGRESSIF"]:
                items = [{"name": k, **v} for k, v in top_data[geo].items()]
                final_tops[geo] = {
                    "by_volume": sorted(items, key=lambda x: x['volume'], reverse=True)[:3],
                    "by_value": sorted(items, key=lambda x: x['value'], reverse=True)[:3],
                    "by_margin": sorted(items, key=lambda x: x['margin'], reverse=True)[:3]
                }

            return {
                "total_sales": total_sales,
                "total_revenue": total_revenue,
                "breakdown": stats,
                "tops": final_tops
            }
            
    except Exception as e: 
        print(f"Stats Error: {e}")
        traceback.print_exc()
        return {}

# --- ROUTES DOSSIERS ---
@app.post("/offers")
def save_offer(offer: OfferRequest):
    if not engine: raise HTTPException(500, "Pas de connexion BDD")
    try:
        with engine.begin() as conn:
            lens_data = offer.lens
            if hasattr(offer, 'correction') and offer.correction:
                 lens_data['correction_data'] = offer.correction
            
            # Sauvegarde avec username et tags
            conn.execute(text("""
                INSERT INTO client_offers (username, encrypted_identity, lens_details, financials, tags) 
                VALUES (:u, :ident, :lens, :fin, :tags)
            """), {
                "u": offer.username,
                "ident": encrypt_dict(offer.client), 
                "lens": json.dumps(lens_data), 
                "fin": json.dumps(offer.finance),
                "tags": json.dumps(offer.tags)
            })
        return {"status": "success"}
    except Exception as e: raise HTTPException(500, str(e))

@app.get("/offers")
def get_offers():
    if not engine: return []
    try:
        with engine.connect() as conn:
            res = conn.execute(text("SELECT * FROM client_offers ORDER BY created_at DESC LIMIT 50"))
            results = []
            for r in res.fetchall():
                try:
                    lens_data = r.lens_details
                    correction = lens_data.get('correction_data', None)
                    results.append({
                        "id": r.id,
                        "date": r.created_at.strftime("%d/%m/%Y %H:%M"),
                        "client": decrypt_dict(r.encrypted_identity),
                        "lens": lens_data,
                        "finance": r.financials,
                        "correction": correction
                    })
                except: continue
            return results
    except: return []

@app.delete("/offers/{offer_id}")
def delete_offer(offer_id: int):
    if not engine: raise HTTPException(500, "Pas de BDD")
    try:
        with engine.begin() as conn:
            result = conn.execute(text("DELETE FROM client_offers WHERE id = :id"), {"id": offer_id})
            if result.rowcount == 0: raise HTTPException(404, "Introuvable")
        return {"status": "success"}
    except Exception as e: raise HTTPException(500, str(e))

# --- ROUTES CATALOGUE ---
@app.get("/lenses")
def get_lenses(type: str = Query(None), brand: str = Query(None), limit: int = Query(3000)):
    if not engine: return []
    try:
        with engine.connect() as conn:
            # On demande tout (*) pour être sûr d'avoir les colonnes bonifiées si elles existent
            sql = f"SELECT * FROM lenses WHERE 1=1"
            params = {}
            if brand: sql += " AND brand ILIKE :brand"; params["brand"] = brand
            if type: 
                type_norm = normalize_string(type)
                if "DEGRESSIF" in type_norm: sql += " AND geometry = 'DEGRESSIF'"
                elif "INTERIEUR" in type_norm: sql += " AND (geometry = 'PROGRESSIF_INTERIEUR' OR geometry = 'INTERIEUR')"
                elif "PROGRESSIF" in type_norm: sql += " AND geometry = 'PROGRESSIF'"
                elif "UNIFOCAL" in type_norm: sql += " AND geometry = 'UNIFOCAL'"
                elif "MULTIFOCAL" in type_norm: sql += " AND geometry = 'MULTIFOCAL'"
                else: sql += " AND geometry ILIKE :geo"; params["geo"] = f"%{type}%"
            
            safe_limit = min(limit, 5000)
            sql += f" ORDER BY purchase_price ASC LIMIT {safe_limit}"
            rows = conn.execute(text(sql), params).fetchall()
            return [row._mapping for row in rows]
    except: return []

@app.post("/upload-catalog")
def upload_catalog(file: UploadFile = File(...)):
    print("🚀 Début requête upload (Threaded)...", flush=True)
    if not engine: raise HTTPException(500, "Serveur BDD déconnecté")
    
    temp_file = f"/tmp/upload_{int(datetime.now().timestamp())}.xlsx"
    print(f"📥 Réception fichier : {file.filename}", flush=True)
    
    try:
        with open(temp_file, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        print("📖 Lecture Excel...", flush=True)
        wb = openpyxl.load_workbook(temp_file, data_only=True, read_only=True)
        
        # MIGRATION : S'assurer que les colonnes bonifiées existent
        with engine.begin() as conn:
            print("♻️  Vidage table & Migration...", flush=True)
            conn.execute(text("ALTER TABLE lenses ADD COLUMN IF NOT EXISTS purchase_price_bonifie DECIMAL(10,2) DEFAULT 0;"))
            conn.execute(text("ALTER TABLE lenses ADD COLUMN IF NOT EXISTS purchase_price_super_bonifie DECIMAL(10,2) DEFAULT 0;"))
            conn.execute(text("TRUNCATE TABLE lenses RESTART IDENTITY CASCADE;"))
        
        total_inserted = 0
        
        with engine.connect() as conn:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_brand = sheet_name.strip().upper()
                print(f"   🔹 Traitement {sheet_brand}...", flush=True)

                row_iterator = sheet.iter_rows(values_only=True)
                header_idx = -1
                headers = []
                
                current_row_idx = 0
                for row in row_iterator:
                    current_row_idx += 1
                    if current_row_idx > 30: break
                    row_str = [str(c).upper() for c in row if c]
                    if any(k in s for s in row_str for k in ["MODELE", "MODÈLE", "LIBELLE", "NAME", "PRIX", "PURCHASE_PRICE"]):
                        headers = row
                        header_idx = current_row_idx
                        break
                
                if not headers: continue

                c_nom = get_col_idx(headers, ['MODELE COMMERCIAL', 'MODELE', 'LIBELLE', 'NAME'])
                c_marque = get_col_idx(headers, ['MARQUE', 'BRAND'])
                c_edi = get_col_idx(headers, ['CODE EDI', 'EDI'])
                c_code = get_col_idx(headers, ['CODE COMMERCIAL', 'COMMERCIAL_CODE'])
                c_geo = get_col_idx(headers, ['GÉOMETRIE', 'GEOMETRIE', 'TYPE', 'GEOMETRY'])
                c_design = get_col_idx(headers, ['DESIGN', 'GAMME'])
                c_idx = get_col_idx(headers, ['INDICE', 'INDEX', 'INDEX_MAT'])
                c_mat = get_col_idx(headers, ['MATIERE', 'MATIÈRE', 'MATERIAL'])
                c_coat = get_col_idx(headers, ['TRAITEMENT', 'COATING'])
                c_flow = get_col_idx(headers, ['FLUX', 'COMMERCIAL_FLOW'])
                c_color = get_col_idx(headers, ['COULEUR', 'COLOR'])
                c_buy = get_col_idx(headers, ['PRIX 2*NETS', 'PRIX', 'ACHAT', 'PURCHASE_PRICE'])
                # Colonnes Alternance
                c_buy_bonif = get_col_idx(headers, ['PRIX BONIFIE', 'ACHAT BONIFIE', 'BONIFIE'])
                c_buy_super = get_col_idx(headers, ['PRIX SUPER BONIFIE', 'SUPER BONIFIE', 'SUPER'])
                
                c_kal = get_col_idx(headers, ['KALIXIA', 'SELL_KALIXIA'])
                c_ite = get_col_idx(headers, ['ITELIS', 'SELL_ITELIS'])
                c_cb = get_col_idx(headers, ['CARTE BLANCHE', 'SELL_CARTEBLANCHE'])
                c_sev = get_col_idx(headers, ['SEVEANE', 'SELL_SEVEANE'])
                c_sant = get_col_idx(headers, ['SANTECLAIRE', 'SANTECLAIR', 'SELL_SANTECLAIR'])

                if c_nom == -1: continue

                batch = []
                BATCH_SIZE = 2000 

                for row in row_iterator:
                    if not row[c_nom]: continue
                    
                    buy = clean_price(row[c_buy]) if c_buy != -1 else 0
                    buy_bonif = clean_price(row[c_buy_bonif]) if c_buy_bonif != -1 else 0
                    buy_super = clean_price(row[c_buy_super]) if c_buy_super != -1 else 0
                    
                    brand = clean_text(row[c_marque]) if c_marque != -1 else sheet_brand
                    if not brand or brand == "None": brand = sheet_brand
                    
                    name = clean_text(row[c_nom])
                    mat = clean_text(row[c_mat]) if c_mat != -1 else ""
                    if any(x in mat.upper() for x in ['TRANS', 'GEN', 'SOLA', 'SUN']): name += f" {mat}"
                    
                    geo_raw = clean_text(row[c_geo]).upper() if c_geo != -1 else ""
                    design_val = clean_text(row[c_design]) if c_design != -1 else "STANDARD"
                    code = clean_text(row[c_code]) if c_code != -1 else ""

                    ltype = 'UNIFOCAL'
                    if 'DEGRESSIF' in geo_raw: ltype = 'DEGRESSIF'
                    elif 'INTERIEUR' in geo_raw: ltype = 'PROGRESSIF_INTERIEUR'
                    elif 'PROG' in geo_raw: ltype = 'PROGRESSIF'
                    elif 'MULTIFOCAL' in geo_raw: ltype = 'MULTIFOCAL'
                    
                    full_search = (name + " " + design_val + " " + code).upper().replace(" ", "")
                    if 'PROXEO' in full_search: ltype = 'DEGRESSIF'
                    if 'MYPROXI' in full_search: ltype = 'PROGRESSIF_INTERIEUR'

                    if buy <= 0:
                         buy = clean_price(row[c_kal]) if c_kal != -1 else 0
                    
                    if buy <= 0: buy = 0.01

                    batch.append({
                        "brand": brand[:100], "edi": clean_text(row[c_edi]) if c_edi != -1 else "",
                        "code": code, "name": name, "geo": ltype, "design": design_val,
                        "idx": clean_index(row[c_idx]) if c_idx != -1 else "1.50",
                        "mat": mat, "coat": clean_text(row[c_coat]) if c_coat != -1 else "DURCI",
                        "flow": clean_text(row[c_flow]) if c_flow != -1 else "FAB",
                        "color": clean_text(row[c_color]) if c_color != -1 else "",
                        "buy": buy, "buy_bonif": buy_bonif, "buy_super": buy_super,
                        "selling": clean_price(row[c_kal]) if c_kal != -1 else 0,
                        "kal": clean_price(row[c_kal]) if c_kal != -1 else 0,
                        "ite": clean_price(row[c_ite]) if c_ite != -1 else 0,
                        "cb": clean_price(row[c_cb]) if c_cb != -1 else 0,
                        "sev": clean_price(row[c_sev]) if c_sev != -1 else 0,
                        "sant": clean_price(row[c_sant]) if c_sant != -1 else 0,
                    })

                    if len(batch) >= BATCH_SIZE:
                        with conn.begin():
                            conn.execute(text("""
                                INSERT INTO lenses (brand, edi_code, commercial_code, name, geometry, design, index_mat, material, coating, commercial_flow, color, purchase_price, purchase_price_bonifie, purchase_price_super_bonifie, selling_price, sell_kalixia, sell_itelis, sell_carteblanche, sell_seveane, sell_santeclair)
                                VALUES (:brand, :edi, :code, :name, :geo, :design, :idx, :mat, :coat, :flow, :color, :buy, :buy_bonif, :buy_super, :selling, :kal, :ite, :cb, :sev, :sant)
                            """), batch)
                        total_inserted += len(batch)
                        batch = []
                        gc.collect()
                        time.sleep(0.01)
                
                if batch:
                    with conn.begin():
                        conn.execute(text("""
                            INSERT INTO lenses (brand, edi_code, commercial_code, name, geometry, design, index_mat, material, coating, commercial_flow, color, purchase_price, purchase_price_bonifie, purchase_price_super_bonifie, selling_price, sell_kalixia, sell_itelis, sell_carteblanche, sell_seveane, sell_santeclair)
                            VALUES (:brand, :edi, :code, :name, :geo, :design, :idx, :mat, :coat, :flow, :color, :buy, :buy_bonif, :buy_super, :selling, :kal, :ite, :cb, :sev, :sant)
                        """), batch)
                    total_inserted += len(batch)
                    gc.collect()

        return {"status": "success", "count": total_inserted}

    except Exception as e:
        print(f"❌ ERREUR: {traceback.format_exc()}", flush=True)
        raise HTTPException(500, f"Erreur: {str(e)}")
    finally:
        if wb: wb.close()
        if os.path.exists(temp_file): os.remove(temp_file)
        gc.collect()