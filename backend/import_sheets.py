import os
import re
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
import openpyxl

# --- CONFIGURATION ---
load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL")

if not DATABASE_URL:
    print("❌ Erreur : DATABASE_URL manquant dans le fichier .env")
    exit()

# --- OUTILS ---
def clean_price(value):
    if not value or value == '' or value == '-': return 0.0
    try:
        clean_s = str(value).replace('€', '').replace('â‚¬', '').replace('%', '').replace(' ', '').replace(',', '.')
        return float(clean_s)
    except ValueError: return 0.0

def clean_index(value):
    if not value: return "1.50"
    match = re.search(r"\d+\.?\d*", str(value).replace(',', '.'))
    return "{:.2f}".format(float(match.group(0))) if match else "1.50"

def clean_text(value):
    return str(value).strip() if value else ""

def get_col_index(headers, candidates):
    """Trouve l'index d'une colonne en cherchant l'un des candidats dans l'en-tête"""
    for i, h in enumerate(headers):
        if h:
            h_clean = str(h).upper().strip()
            for c in candidates:
                if c.upper() in h_clean: # 'in' permet de matcher "sell_kalixia" avec "KALIXIA"
                    return i
    return -1

def import_data_from_excel():
    print("🚀 Démarrage Importation (Mode ROBUSTE - TEXT)...")
    
    excel_file = "catalogue.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Erreur : Le fichier '{excel_file}' est introuvable.")
        return

    engine = create_engine(DATABASE_URL)
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        with engine.begin() as conn:
            print("🧹 Suppression de l'ancienne table 'lenses'...")
            # On force la suppression pour recréer avec la bonne structure
            conn.execute(text("DROP TABLE IF EXISTS lenses CASCADE;"))
            
            print("🏗️ Création de la nouvelle table (Types TEXT illimités)...")
            # Utilisation de TEXT au lieu de VARCHAR pour éviter les erreurs de longueur
            conn.execute(text("""
                CREATE TABLE lenses (
                    id SERIAL PRIMARY KEY,
                    brand VARCHAR(100),
                    edi_code TEXT,                 -- Changé en TEXT
                    commercial_code TEXT,          -- Changé en TEXT
                    name TEXT,                     -- Changé en TEXT
                    geometry VARCHAR(100),
                    design TEXT,                   -- Changé en TEXT
                    index_mat VARCHAR(50),
                    material TEXT,                 -- Changé en TEXT
                    coating TEXT,                  -- Changé en TEXT
                    commercial_flow VARCHAR(100),
                    color VARCHAR(100),
                    purchase_price DECIMAL(10,2), selling_price DECIMAL(10,2),
                    sell_kalixia DECIMAL(10,2), sell_itelis DECIMAL(10,2),
                    sell_carteblanche DECIMAL(10,2), sell_seveane DECIMAL(10,2),
                    sell_santeclair DECIMAL(10,2)
                );
            """))

            stmt = text("""
                INSERT INTO lenses (
                    brand, edi_code, commercial_code, name, geometry, design, index_mat, material, coating, 
                    commercial_flow, color, purchase_price, selling_price, 
                    sell_kalixia, sell_itelis, sell_carteblanche, sell_seveane, sell_santeclair
                ) VALUES (
                    :brand, :edi, :code, :name, :geo, :design, :idx, :mat, :coat, 
                    :flow, :color, :buy, :selling, 
                    :p_kal, :p_ite, :p_cb, :p_sev, :p_sant
                )
            """)

            total_count = 0
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                current_brand = sheet_name.strip().upper()
                print(f"\n📄 Analyse Feuille : {current_brand}")
                
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    print("   ⚠️ Feuille vide.")
                    continue
                
                # --- RECHERCHE INTELLIGENTE DE L'EN-TÊTE ---
                header_row_idx = -1
                headers = []
                
                # On cherche une ligne qui contient des mots clés
                keywords = ["MODELE", "MARQUE", "BRAND", "NAME", "GEOMETRY"]
                
                for idx, row in enumerate(rows[:10]): 
                    row_str = [str(c).upper() for c in row if c]
                    if any(k in s for s in row_str for k in keywords):
                        header_row_idx = idx
                        headers = row
                        print(f"   ✅ En-têtes trouvés à la ligne {idx + 1}")
                        break
                
                if header_row_idx == -1:
                    print("   ❌ Impossible de trouver la ligne d'en-tête.")
                    continue

                # --- MAPPING DES COLONNES ---
                c_marque = get_col_index(headers, ['MARQUE', 'FABRICANT', 'BRAND'])
                c_name = get_col_index(headers, ['MODELE COMMERCIAL', 'MODELE', 'LIBELLE', 'NAME'])
                c_buy = get_col_index(headers, ['PRIX 2*NETS', '2*NETS', 'ACHAT', 'PURCHASE_PRICE'])
                
                if c_name == -1:
                    print("   ❌ Colonne 'MODELE' introuvable.")
                    continue
                
                # Mapping des autres colonnes
                c_edi = get_col_index(headers, ['CODE EDI', 'EDI_CODE'])
                c_code = get_col_index(headers, ['CODE COMMERCIAL', 'COMMERCIAL_CODE'])
                c_geo = get_col_index(headers, ['GÉOMETRIE', 'GEOMETRIE', 'TYPE', 'GEOMETRY'])
                c_design = get_col_index(headers, ['DESIGN', 'GAMME'])
                c_idx = get_col_index(headers, ['INDICE', 'INDEX_MAT'])
                c_mat = get_col_index(headers, ['MATIERE', 'MATERIAL'])
                c_coat = get_col_index(headers, ['TRAITEMENT', 'COATING'])
                c_flow = get_col_index(headers, ['FLUX COMMERCIAL', 'FLUX', 'COMMERCIAL_FLOW'])
                c_color = get_col_index(headers, ['COULEUR', 'COLOR'])
                
                # Prix Réseaux
                c_kal = get_col_index(headers, ['KALIXIA', 'SELL_KALIXIA'])
                c_ite = get_col_index(headers, ['ITELIS', 'SELL_ITELIS'])
                c_cb = get_col_index(headers, ['CARTE BLANCHE', 'SELL_CARTEBLANCHE'])
                c_sev = get_col_index(headers, ['SEVEANE', 'SELL_SEVEANE'])
                c_sant = get_col_index(headers, ['SANTECLAIRE', 'SANTECLAIR', 'SELL_SANTECLAIR'])

                sheet_inserted = 0
                
                for row in rows[header_row_idx + 1:]:
                    if not row[c_name]: continue
                    
                    brand = clean_text(row[c_marque]) if c_marque != -1 else current_brand
                    if not brand or brand == "None": brand = current_brand
                    
                    raw_name = clean_text(row[c_name])
                    raw_geo = clean_text(row[c_geo]).upper() if c_geo != -1 else ""
                    
                    ltype = 'UNIFOCAL'
                    if 'PROG' in raw_geo: ltype = 'PROGRESSIF'
                    elif 'DEGRESSIF' in raw_geo or 'INTERIEUR' in raw_geo: ltype = 'DEGRESSIF'
                    elif 'MULTIFOCAL' in raw_geo: ltype = 'MULTIFOCAL'
                    
                    purchase = clean_price(row[c_buy]) if c_buy != -1 else 0
                    
                    if purchase == 0 and c_buy != -1:
                         continue 

                    # On coupe les textes trop longs pour brand au cas où, mais les autres sont en TEXT
                    safe_brand = brand[:100]

                    params = {
                        "brand": safe_brand,
                        "edi": clean_text(row[c_edi]) if c_edi != -1 else "",
                        "code": clean_text(row[c_code]) if c_code != -1 else "",
                        "name": raw_name,
                        "geo": ltype,
                        "design": clean_text(row[c_design]) if c_design != -1 else "STANDARD",
                        "idx": clean_index(row[c_idx]) if c_idx != -1 else "1.50",
                        "mat": clean_text(row[c_mat]) if c_mat != -1 else "",
                        "coat": clean_text(row[c_coat]) if c_coat != -1 else "DURCI",
                        "flow": clean_text(row[c_flow]) if c_flow != -1 else "FAB",
                        "color": clean_text(row[c_color]) if c_color != -1 else "",
                        "buy": purchase,
                        "selling": clean_price(row[c_kal]) if c_kal != -1 else 0, 
                        "p_kal": clean_price(row[c_kal]) if c_kal != -1 else 0,
                        "p_ite": clean_price(row[c_ite]) if c_ite != -1 else 0,
                        "p_cb": clean_price(row[c_cb]) if c_cb != -1 else 0,
                        "p_sev": clean_price(row[c_sev]) if c_sev != -1 else 0,
                        "p_sant": clean_price(row[c_sant]) if c_sant != -1 else 0,
                    }
                    
                    conn.execute(stmt, params)
                    sheet_inserted += 1
                    total_count += 1

                print(f"      -> {sheet_inserted} verres insérés.")

            print(f"\n🎉 TERMINE : {total_count} verres importés au total.")

    except Exception as e:
        print(f"❌ ERREUR CRITIQUE : {e}")

if __name__ == "__main__":
    import_data_from_excel()