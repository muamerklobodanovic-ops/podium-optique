import os
import re
import json
import time
from datetime import datetime
from dotenv import load_dotenv
from sqlalchemy import create_engine, text
import openpyxl
import sys

# 1. Configuration
load_dotenv()
DATABASE_URL = os.getenv("DATABASE_URL")

if not DATABASE_URL:
    print("❌ Erreur : DATABASE_URL manquant dans le fichier .env")
    sys.exit(1)

if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
if "sslmode" not in DATABASE_URL:
    separator = "&" if "?" in DATABASE_URL else "?"
    DATABASE_URL += f"{separator}sslmode=require"

try:
    engine = create_engine(DATABASE_URL, echo=False)
    print("✅ Connexion BDD établie.")
except Exception as e:
    print(f"❌ Erreur de connexion : {e}")
    sys.exit(1)

# --- OUTILS ---
def clean_price(value):
    if not value or value == '' or value == '-': return 0.0
    try:
        s = str(value).replace('€', '').replace('â‚¬', '').replace('%', '').replace(' ', '').replace('\xa0', '').strip()
        s = s.replace(',', '.')
        return float(s)
    except: return 0.0

def clean_index(value):
    if not value: return "1.50"
    match = re.search(r"\d+\.?\d*", str(value).replace(',', '.'))
    return "{:.2f}".format(float(match.group(0))) if match else "1.50"

def clean_text(value): return str(value).strip() if value else ""

def get_col_idx(headers, candidates):
    for i, h in enumerate(headers):
        if h:
            h_str = str(h).upper().strip()
            if any(c.upper() in h_str for c in candidates): 
                return i
    return -1

def import_data():
    excel_file = "catalogue.xlsx"
    if not os.path.exists(excel_file):
        print(f"❌ Fichier '{excel_file}' introuvable.")
        return

    print(f"🚀 Démarrage DIAGNOSTIC de '{excel_file}'...")
    start_time = time.time()

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
        
        with engine.begin() as conn:
            print("♻️  Réinitialisation de la table 'lenses'...")
            conn.execute(text("DROP TABLE IF EXISTS lenses CASCADE;"))
            conn.execute(text("""
                CREATE TABLE lenses (
                    id SERIAL PRIMARY KEY,
                    brand TEXT, edi_code TEXT, commercial_code TEXT, name TEXT,
                    geometry TEXT, design TEXT, index_mat TEXT,
                    material TEXT, coating TEXT, commercial_flow TEXT, color TEXT,
                    purchase_price DECIMAL(10,2), selling_price DECIMAL(10,2),
                    sell_kalixia DECIMAL(10,2), sell_itelis DECIMAL(10,2),
                    sell_carteblanche DECIMAL(10,2), sell_seveane DECIMAL(10,2),
                    sell_santeclair DECIMAL(10,2)
                );
            """))

        total_inserted = 0

        with engine.connect() as conn:
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_brand = sheet_name.strip().upper()
                print(f"\n🔹 Feuille : {sheet_brand}")

                row_iterator = sheet.iter_rows(values_only=True)
                
                header_idx = -1
                headers = []
                current_idx = 0
                
                # Recherche En-tête
                for row in row_iterator:
                    current_idx += 1
                    if current_idx > 30: break
                    row_str = [str(c).upper() for c in row if c]
                    if any(k in s for s in row_str for k in ["MODELE", "MODÈLE", "LIBELLE", "NAME", "PRIX"]):
                        headers = row
                        header_idx = current_idx
                        print(f"   ✅ En-têtes trouvés ligne {current_idx}")
                        print(f"      Aperçu en-têtes: {[str(h)[:15] for h in row if h][:6]}...")
                        break
                
                if not headers:
                    print(f"   ⚠️ Pas d'en-tête trouvé, feuille ignorée.")
                    continue

                # Mapping et Diagnostic Colonnes
                c_nom = get_col_idx(headers, ['MODELE COMMERCIAL', 'MODELE', 'LIBELLE', 'NAME'])
                c_buy = get_col_idx(headers, ['PRIX 2*NETS', 'PRIX', 'ACHAT'])
                
                if c_nom == -1: 
                    print("   ❌ Colonne 'MODÈLE' introuvable dans les en-têtes !")
                    continue
                
                if c_buy == -1:
                    print("   ⚠️ Colonne 'PRIX ACHAT' introuvable ! Les verres seront ignorés (Sécurité).")
                    # On continue pour voir si on trouve autre chose, mais ça n'insérera rien si logic buy<=0
                
                c_marque = get_col_idx(headers, ['MARQUE', 'BRAND'])
                c_edi = get_col_idx(headers, ['CODE EDI', 'EDI'])
                c_code = get_col_idx(headers, ['CODE COMMERCIAL', 'COMMERCIAL_CODE'])
                c_geo = get_col_idx(headers, ['GÉOMETRIE', 'GEOMETRIE', 'TYPE'])
                c_design = get_col_idx(headers, ['DESIGN', 'GAMME'])
                c_idx = get_col_idx(headers, ['INDICE', 'INDEX'])
                c_mat = get_col_idx(headers, ['MATIERE', 'MATIÈRE'])
                c_coat = get_col_idx(headers, ['TRAITEMENT', 'COATING'])
                c_flow = get_col_idx(headers, ['FLUX'])
                c_color = get_col_idx(headers, ['COULEUR'])
                
                c_kal = get_col_idx(headers, ['KALIXIA'])
                c_ite = get_col_idx(headers, ['ITELIS'])
                c_cb = get_col_idx(headers, ['CARTE BLANCHE'])
                c_sev = get_col_idx(headers, ['SEVEANE'])
                c_sant = get_col_idx(headers, ['SANTECLAIRE'])

                print(f"   ℹ️ Colonnes identifiées -> Modèle: {c_nom}, Prix Achat: {c_buy}, Géométrie: {c_geo}, Indice: {c_idx}")

                batch = []
                BATCH_SIZE = 1000
                skipped_count = 0

                for row in row_iterator:
                    # Sécurité lecture ligne vide
                    if not row or len(row) <= c_nom or not row[c_nom]: 
                        continue
                    
                    # Analyse Prix
                    buy = 0
                    if c_buy != -1 and len(row) > c_buy:
                        buy = clean_price(row[c_buy])
                    
                    # DEBUG PREMIERE LIGNE REJETÉE
                    if buy <= 0 and skipped_count == 0:
                        print(f"      ⚠️ Première ligne rejetée (Prix nul ou non trouvé) : {row[c_nom]}")
                        skipped_count += 1
                        continue
                    if buy <= 0:
                        skipped_count += 1
                        continue

                    brand = clean_text(row[c_marque]) if c_marque != -1 and len(row) > c_marque else sheet_brand
                    if not brand or brand == "None": brand = sheet_brand
                    
                    name = clean_text(row[c_nom])
                    mat = clean_text(row[c_mat]) if c_mat != -1 and len(row) > c_mat else ""
                    if any(x in mat.upper() for x in ['TRANS', 'GEN', 'SOLA']): name += f" {mat}"
                    
                    geo_raw = clean_text(row[c_geo]).upper() if c_geo != -1 and len(row) > c_geo else ""
                    ltype = 'UNIFOCAL'
                    if 'PROG' in geo_raw: ltype = 'PROGRESSIF'
                    elif 'DEGRESSIF' in geo_raw: ltype = 'DEGRESSIF'
                    elif 'MULTIFOCAL' in geo_raw: ltype = 'MULTIFOCAL'

                    lens = {
                        "brand": brand[:100], 
                        "edi": clean_text(row[c_edi]) if c_edi != -1 and len(row) > c_edi else "",
                        "code": clean_text(row[c_code]) if c_code != -1 and len(row) > c_code else "",
                        "name": name,
                        "geo": ltype,
                        "design": clean_text(row[c_design]) if c_design != -1 and len(row) > c_design else "STANDARD",
                        "idx": clean_index(row[c_idx]) if c_idx != -1 and len(row) > c_idx else "1.50",
                        "mat": mat,
                        "coat": clean_text(row[c_coat]) if c_coat != -1 and len(row) > c_coat else "DURCI",
                        "flow": clean_text(row[c_flow]) if c_flow != -1 and len(row) > c_flow else "FAB",
                        "color": clean_text(row[c_color]) if c_color != -1 and len(row) > c_color else "",
                        "buy": buy,
                        "selling": clean_price(row[c_kal]) if c_kal != -1 and len(row) > c_kal else 0,
                        "kal": clean_price(row[c_kal]) if c_kal != -1 and len(row) > c_kal else 0,
                        "ite": clean_price(row[c_ite]) if c_ite != -1 and len(row) > c_ite else 0,
                        "cb": clean_price(row[c_cb]) if c_cb != -1 and len(row) > c_cb else 0,
                        "sev": clean_price(row[c_sev]) if c_sev != -1 and len(row) > c_sev else 0,
                        "sant": clean_price(row[c_sant]) if c_sant != -1 and len(row) > c_sant else 0,
                    }
                    batch.append(lens)

                    if len(batch) >= BATCH_SIZE:
                        with conn.begin():
                            conn.execute(text("""
                                INSERT INTO lenses (brand, edi_code, commercial_code, name, geometry, design, index_mat, material, coating, commercial_flow, color, purchase_price, selling_price, sell_kalixia, sell_itelis, sell_carteblanche, sell_seveane, sell_santeclair)
                                VALUES (:brand, :edi, :code, :name, :geo, :design, :idx, :mat, :coat, :flow, :color, :buy, :selling, :kal, :ite, :cb, :sev, :sant)
                            """), batch)
                        total_inserted += len(batch)
                        print(f"   ... {total_inserted} verres insérés", end="\r")
                        batch = []
                
                if batch:
                    with conn.begin():
                        conn.execute(text("""
                            INSERT INTO lenses (brand, edi_code, commercial_code, name, geometry, design, index_mat, material, coating, commercial_flow, color, purchase_price, selling_price, sell_kalixia, sell_itelis, sell_carteblanche, sell_seveane, sell_santeclair)
                            VALUES (:brand, :edi, :code, :name, :geo, :design, :idx, :mat, :coat, :flow, :color, :buy, :selling, :kal, :ite, :cb, :sev, :sant)
                        """), batch)
                    total_inserted += len(batch)
                
                if skipped_count > 0:
                    print(f"      ⚠️ {skipped_count} lignes ignorées (Prix nul ou vide).")

        
        end_time = time.time()
        print(f"\n\n✅ SUCCÈS ! {total_inserted} verres importés en {round(end_time - start_time, 2)} secondes.")

    except Exception as e:
        print(f"\n❌ ERREUR CRITIQUE : {e}")

if __name__ == "__main__":
    import_data()